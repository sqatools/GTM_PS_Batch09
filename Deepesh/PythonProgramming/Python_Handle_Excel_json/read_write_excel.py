import openpyxl
# pip install openpyxl

def read_excel_data(filepath, sheet_name, cell_name):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb[sheet_name]
    cell = sheet[cell_name]
    print(cell.value)

# read_excel_data("test_data.xlsx", "Sheet1", "B1")
# India
#read_excel_data("test_data.xlsx", "Sheet2", "A1")
# Python Programming

# for i in range(1, 6):
#     read_excel_data("test_data.xlsx", "Sheet1", f"A{i}")

# Read three value from first colum or each sheet
"""
for i in range(1, 6):
    print(f"sheet name :Sheet{i}", )
    for j in range(1, 4):

         read_excel_data("test_data.xlsx", f"Sheet{i}", f"A{j}")

    print("_"*50)
"""
# Write a program to read first value for each 5 colum from each sheet.

for i in range(1, 6):
    print(f"sheet name :Sheet{i}", )
    for j in ['A', 'B', 'C', 'D', 'E']:

         read_excel_data("test_data.xlsx", f"Sheet{i}", f"{j}1")
         read_excel_data("test_data.xlsx", f"Sheet{i}", f"{j}2")

    print("_"*50)



# for i in range(1, 6):
#     for j in range(1, 4):
#          read_excel_data("test_data.xlsx", f"Sheet{i}", f"A{j}")
#
#     print("_"*50)

# write content to excel sheet
def write_excel_data(filename, sheet_name, cell_name, cell_val):

    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheet_name]
    cell_n = sheet[cell_name]
    cell_n.value = cell_val
    wb.save(filename)

#write_excel_data("test_data.xlsx", "Sheet3", "A1", "Indore")
# replace exiting data with blank.
#write_excel_data("test_data.xlsx", "Sheet3", "A1", "")
# read all data of specific sheet


def read_all_data(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    max_row = sheet.max_row
    print("Max row :", max_row)
    max_cols = sheet.max_column
    print("Max colns :", max_cols)
    # number of rows
    for i in range(1, max_row+1):
        # number of columns
        for j in range(1, max_cols+1):
            cell_value = sheet.cell(i, j)
            print(cell_value.value, end=" ")

        print()

    print(max_cols, max_row)


print("_"*50)

read_all_data("test_data.xlsx", "Sheet4")

# Home Work
# write a python program to read first three cell of 5 different sheet.
# Write a python to copy the content one column from one sheet to another.

