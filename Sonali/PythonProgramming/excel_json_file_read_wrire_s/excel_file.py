import openpyxl                        #importing the excel file, so that excel can be used via python



#print the values from the excel sheet

def excel_data(filepath, sheet_name, cell_name):
    wb = openpyxl.load_workbook(filepath)
    sheet9 = wb[sheet_name]

    cell = sheet9[cell_name]
    print(sheet.cell(i,j).value)
    print(cell.value)



excel_data("python_sonali.xlsx","Sheet1","A1")

excel_data("python_sonali.xlsx","Sheet2","B1")   # nothing is there in that cell

for i in range(1, 6):
    excel_data("python_sonali.xlsx","Sheet1", f"A{i}")



print("*"*90)

Write a python to copy the content one column from one sheet to another

wb1=openpyxl.load_workbook("python_sonali.xlsx")   #load a excel file
sheet1=wb1['Sheet4']                #both source sheet and destination sheet define
sheet2=wb1['Sheet5']
maxr=sheet1.max_row                #the sheet u want to copy the sheet max row and max col define
maxc=sheet1.max_column
print("Max row :", maxr)
print("Max col :", maxc)
for i in range(1,maxr+1):
    for j in range(1,maxc+1):
        sheet2.cell(i,j).value=sheet1.cell(i,j).value  #within range max row and max col copy the source sheet to destination sheet
wb1.save("python_sonali.xlsx")                       #save the value


#sheet2.cell(i,j).value=sheet1.cell(i,j).value ------------this means to sheet2 i, j meand row and col we are coping sheet1 row and col value


print("*"*80)

write a python program to read first three cell of 5 different sheet.

def three_cell(filepath,sheet_name,cell_name):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb[sheet_name]
    cell = sheet[cell_name]
    print(cell.value)
for i in range(1, 6):
    print(f"sheet name:Sheet{i}")
    for j in range(1, 4):
        three_cell("python_sonali.xlsx", f"Sheet{i}", f"A{j}")


write the python code to read multiple rows and  multiple columns from the sheet

def mult_cell_values(filename,sheet_name,cell_name):
    wb3 = openpyxl.load_workbook(filename)
    sheet = wb3[sheet_name]
    cell = sheet[cell_name]
    print(cell.value)
for i in range(1,6):
    print(f"sheet name:Sheet{i}")
    for j in ['A','B','C','D']:

        mult_cell_values("python_sonali.xlsx", f"Sheet{i}", f"{j}1")
        mult_cell_values("python_sonali.xlsx", f"Sheet{i}", f"{j}2")
        mult_cell_values("python_sonali.xlsx", f"Sheet{i}", f"{j}3")
























