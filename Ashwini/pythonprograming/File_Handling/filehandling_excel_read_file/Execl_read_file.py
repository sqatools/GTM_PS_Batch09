import openpyxl
# pip install openpyxl

def read_excel_data(filepath):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb['Sheet1']
    cell = sheet['A1']
    print(cell.value)

read_excel_data("test_data.xlsx")
print("-"*50)
def read_excel_data(filepath,Sheet_name,cell_name):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb[Sheet_name]
    cell = sheet[cell_name]
    print(cell.value)

read_excel_data("test_data.xlsx", "Sheet1", "B1")
read_excel_data("test_data.xlsx", "Sheet2", "A1")
print("-"*50)
#for i in range(1, 6):
 #    read_excel_data("test_data.xlsx", "Sheet1", f"A{i}")


# write content to excel sheet
def write_excel_data(filename, sheet_name, cell_name, cell_val):

    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheet_name]
    cell_n = sheet[cell_name]
    cell_n.value = cell_val
    wb.save(filename)

write_excel_data("test_data.xlsx", "Sheet6", "A1", "Indore")
write_excel_data("test_data.xlsx", "Sheet6", "B1", "India")
# replace exiting data with blank.
#write_excel_data("test_data.xlsx", "Sheet3", "A1", "")
# read all data of specific sheet
def read_all_data(filepath,sheet_name):
    wb=openpyxl.load_workbook(filepath)
    sheet=wb[sheet_name]
    max_row=sheet.max_row
    max_cols=sheet.max_column
    for i in range(1,max_row):
        for j in range(1,max_cols):
            cell_value=sheet.cell(i,j)
            print(cell_value.value,end=" ")
read_all_data("test_data.xlsx","Sheet3")


print("-"*50)
# Home Work
# write a python program to read first three cell of 5 different sheet.
"""#Load the workbook
workbook = openpyxl.load_workbook('test_data.xlsx')

# List of sheet names to read from
sheet_names = ['Sheet1', 'Sheet2', 'Sheet3', 'Sheet4', 'Sheet5']

# Dictionary to store the cell values
data = {}

# Iterate through each sheet
for sheet_name in sheet_names:
    sheet = workbook[sheet_name]
    # Read the first three cells (A1, B1, C1)
    cell_values = {
        'A1': sheet['A1'].value,
        'B1': sheet['B1'].value,
        'C1': sheet['C1'].value
    }
    data[sheet_name] = cell_values

# Print the cell values
for sheet_name, cell_values in data.items():
    print(f"Sheet: {sheet_name}")
    for cell, value in cell_values.items():
        print(f"{cell}: {value}")
    print()
"""
def read_first_three_excel_data(filepath,Sheet_name,cell_name):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb[Sheet_name]
    cell = sheet[cell_name]
    print(cell.value)
for i in range(1,6):
    print(f"sheet_name:Sheet{i}",)
    for j in range(1,4):
        read_first_three_excel_data("test_data.xlsx", f"Sheet{i}", f"A{j}")
    print()

# Write a python to copy the content one column from one sheet to another.
def copy_clo_sheet(filepath):
    wb=openpyxl.load_workbook(filepath)
    source_sheet = wb['Sheet1']
    destination_sheet = wb['Sheet7']
    source_column = 'A'
    for row in range(1, source_sheet.max_row + 1):
        cell_value = source_sheet[f'{source_column}{row}'].value
        destination_sheet[f'{source_column}{row}'].value = cell_value
        wb.save(filepath)  
    print("Column copied successfully!")
copy_clo_sheet("test_data.xlsx")

