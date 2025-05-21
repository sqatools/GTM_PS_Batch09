# write a python program to read first three cell of 5 different sheet.
import openpyxl


def read_data_excel(filepath, sheet_name, cell_name):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb[sheet_name]
    cell = sheet[cell_name]
    print(cell.value)


for i in range(1,6):#here i is sheets
    for j in range(1,4):#here j is no of rows
        read_data_excel("test_data.xlsx", f"Sheet{i}", f"A{j}")

#read_data_excel("test_data.xlsx", "Sheet1", "A1")
print("_"*50)




# Write a python to copy the content one column from one sheet to another.
def R_read_data_excel(filepath, R_sheet_name, R_cell_name):
    wb = openpyxl.load_workbook(filepath)
    R_sheet = wb[R_sheet_name]
    R_cell = R_sheet[R_cell_name]
    print(R_cell.value)

def write_data_excel(filepath, sheet_name, cell_name, cell_value):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb[sheet_name]
    cell_n = sheet[cell_name]
    cell_n.value = cell_value
    wb.save(filepath)


R_read_data_excel("test_data.xlsx","Sheet1","A1")
write_data_excel("test_data.xlsx","Sheet6","A1",)

